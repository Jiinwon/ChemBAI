#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
DART_MTL 예측 (wide format) + fingerprint 디스크 캐시 재사용 + data_prob/data_pred 분리

입력:
  .../prediction/DART_MTL_251228/DART_MTL_seed{seed}{split}/seed{seed}{split}.xlsx
  - sheet: "data"
  - columns: SMILES + assay columns

모델 탐색(우선순위):
  1) .../training/DART_MTL/results/seed_{seed}/{assay}_{model}_{fp}/model.joblib
  2) 없으면 .../training/DART_MTL/results/backup/seed_{seed}/{assay}_{model}_{fp}/model.joblib

fingerprints 캐시:
  .../DART_MTL_seed{seed}{split}/fingerprints/FP_{fp}_dim{dim}.npz
  - X: uint8 (0/1) shape (N, dim)
  - valid_mask: uint8 (0/1) shape (N,)

출력(조합별 1개):
  results/{model}_{fp}/seed{seed}{split}_{model}_{fp}.xlsx
  - truth_data : 원본 data(정답 포함) 보존
  - data_prob  : SMILES + assay 전체, 값=예측확률
  - data_pred  : SMILES + assay 전체, 값=0/1 라벨 (threshold 기준)
  - pred_info  : assay별 상태/차원/모델 경로(실사용 경로) 기록

점수:
  score_seed{seed}{split}.xlsx
  columns: Assay Name, Model, MF Metric, AUC, accuracy, precision, recall, F1, n_eval

평가:
  - 정답이 0/1인 행만 포함
  - 예측이 NaN인 행 제외
  - AUC는 y_true가 0/1 두 클래스 모두 있을 때만 계산
"""

import argparse
import shutil
from collections import OrderedDict
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd
from joblib import load as joblib_load

from rdkit import Chem, DataStructs
from rdkit.Chem import AllChem, MACCSkeys

from sklearn.metrics import (
    roc_auc_score,
    accuracy_score,
    precision_score,
    recall_score,
    f1_score,
)

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


# =========================
# 고정 설정
# =========================
MODELS = ["dt", "rf", "xgb", "gbt", "logistic"]
FINGERPRINTS = ["MACCS", "Morgan", "RDKit", "Pattern", "Layered"]

TOXCAST_MODEL_ROOT = Path(
    "/home1/won0316/_RESEARCH/0817_Genotoxicity/1_Git_upload/ChemBAI/ToxCast_model"
)
TRAIN_RESULTS_ROOT = TOXCAST_MODEL_ROOT / "experiments/training/DART_MTL/results"
TRAIN_BACKUP_ROOT = TRAIN_RESULTS_ROOT / "backup"

PRED_BASE_ROOT = TOXCAST_MODEL_ROOT / "experiments/prediction/DART_MTL_251228"

EXCEL_ENGINE = "openpyxl"
MACCS_BITS = 167


# =========================
# 유틸: 컬럼 찾기
# =========================
def find_smiles_col(columns: List[str]) -> str:
    for c in columns:
        if str(c).strip().lower() == "smiles":
            return c
    raise ValueError("data 시트에서 'SMILES' 열을 찾지 못했습니다.")


# =========================
# 유틸: 모델에서 기대 feature 수 / feature name 추출
# =========================
def get_n_features_in(model) -> Optional[int]:
    if hasattr(model, "n_features_in_"):
        return int(model.n_features_in_)
    if hasattr(model, "steps"):
        for _, step in reversed(model.steps):
            if hasattr(step, "n_features_in_"):
                return int(step.n_features_in_)
    return None


def get_feature_names_in(model) -> Optional[np.ndarray]:
    if hasattr(model, "feature_names_in_"):
        return np.array(model.feature_names_in_)
    if hasattr(model, "steps"):
        for _, step in reversed(model.steps):
            if hasattr(step, "feature_names_in_"):
                return np.array(step.feature_names_in_)
    return None


def ensure_feature_named_input(model, X_np: np.ndarray) -> Union[np.ndarray, pd.DataFrame]:
    names = get_feature_names_in(model)
    if names is not None and len(names) == X_np.shape[1]:
        return pd.DataFrame(X_np, columns=names)
    return X_np


# =========================
# 유틸: FP 생성
# =========================
def smiles_to_mol(smiles: str) -> Optional[Chem.Mol]:
    if smiles is None:
        return None
    s = str(smiles).strip()
    if s == "" or s.lower() == "nan":
        return None
    return Chem.MolFromSmiles(s)


def bitvect_to_uint8(fp, dim: int) -> np.ndarray:
    arr = np.zeros((dim,), dtype=np.int8)
    DataStructs.ConvertToNumpyArray(fp, arr)
    return arr.astype(np.uint8, copy=False)


def compute_fingerprint_matrix_uint8(
    smiles_list: List[str],
    fp_name: str,
    dim: int,
) -> Tuple[np.ndarray, np.ndarray]:
    n = len(smiles_list)
    X = np.zeros((n, dim), dtype=np.uint8)
    valid_mask = np.zeros((n,), dtype=bool)

    for i, smi in enumerate(smiles_list):
        mol = smiles_to_mol(smi)
        if mol is None:
            continue

        if fp_name == "MACCS":
            fp = MACCSkeys.GenMACCSKeys(mol)
        elif fp_name == "Morgan":
            fp = AllChem.GetMorganFingerprintAsBitVect(mol, radius=2, nBits=dim)
        elif fp_name == "RDKit":
            fp = Chem.RDKFingerprint(mol, fpSize=dim)
        elif fp_name == "Pattern":
            fp = Chem.PatternFingerprint(mol, fpSize=dim)
        elif fp_name == "Layered":
            fp = Chem.LayeredFingerprint(mol, fpSize=dim)
        else:
            raise ValueError(f"Unknown fingerprint: {fp_name}")

        if fp.GetNumBits() != dim:
            continue

        valid_mask[i] = True
        X[i, :] = bitvect_to_uint8(fp, dim)

    return X, valid_mask


def safe_predict_proba(model, X_in) -> np.ndarray:
    if hasattr(model, "predict_proba"):
        proba = model.predict_proba(X_in)
        if proba.ndim == 2 and proba.shape[1] >= 2:
            return proba[:, 1].astype(np.float64)
        return np.asarray(proba, dtype=np.float64).reshape(-1)

    if hasattr(model, "decision_function"):
        score = model.decision_function(X_in).astype(np.float64)
        return 1.0 / (1.0 + np.exp(-score))

    pred = model.predict(X_in)
    return np.asarray(pred, dtype=np.float64).reshape(-1)


# =========================
# 점수 계산
# =========================
def compute_metrics(y_true: np.ndarray, y_proba: np.ndarray, thr: float) -> Dict[str, float]:
    out = {"AUC": np.nan, "accuracy": np.nan, "precision": np.nan, "recall": np.nan, "F1": np.nan}
    y_pred = (y_proba >= thr).astype(int)

    if np.unique(y_true).size >= 2:
        out["AUC"] = float(roc_auc_score(y_true, y_proba))

    out["accuracy"] = float(accuracy_score(y_true, y_pred))
    out["precision"] = float(precision_score(y_true, y_pred, zero_division=0))
    out["recall"] = float(recall_score(y_true, y_pred, zero_division=0))
    out["F1"] = float(f1_score(y_true, y_pred, zero_division=0))
    return out


# =========================
# 엑셀 작성 (data_prob / data_pred)
# =========================
def replace_or_create_sheet(wb: openpyxl.Workbook, name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if name in wb.sheetnames:
        wb.remove(wb[name])
    return wb.create_sheet(name)


def write_df_sheet(wb: openpyxl.Workbook, df: pd.DataFrame, name: str):
    ws = replace_or_create_sheet(wb, name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)


def write_pred_info_sheet(wb: openpyxl.Workbook, rows: List[Dict[str, object]], name: str = "pred_info"):
    ws = replace_or_create_sheet(wb, name)
    if not rows:
        ws.append(["info", "no rows"])
        return
    cols = list(rows[0].keys())
    ws.append(cols)
    for row in rows:
        ws.append([row.get(c) for c in cols])


# =========================
# FP 디스크 캐시
# =========================
def fp_cache_path(fp_dir: Path, fp_name: str, dim: int) -> Path:
    return fp_dir / f"FP_{fp_name}_dim{dim}.npz"


def load_fp_from_disk(fp_path: Path) -> Tuple[np.ndarray, np.ndarray]:
    z = np.load(fp_path, allow_pickle=False)
    X = z["X"]  # uint8
    valid_u8 = z["valid_mask"]  # uint8
    valid_mask = valid_u8.astype(bool)
    return X, valid_mask


def save_fp_to_disk(fp_path: Path, X_uint8: np.ndarray, valid_mask: np.ndarray):
    fp_path.parent.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(
        fp_path,
        X=X_uint8.astype(np.uint8, copy=False),
        valid_mask=valid_mask.astype(np.uint8, copy=False),
    )


class LRUCache:
    """
    fp (fp_name, dim, mem_dtype) -> (X_mem, valid_mask)
    """
    def __init__(self, max_items: int = 4):
        self.max_items = max_items
        self.od = OrderedDict()

    def get(self, key):
        if key not in self.od:
            return None
        self.od.move_to_end(key)
        return self.od[key]

    def put(self, key, value):
        self.od[key] = value
        self.od.move_to_end(key)
        while len(self.od) > self.max_items:
            self.od.popitem(last=False)


def load_or_build_fp(
    smiles_list: List[str],
    fp_name: str,
    dim: int,
    fp_dir: Path,
    mem_cache: LRUCache,
    mem_dtype: str,
    rebuild: bool = False,
) -> Tuple[np.ndarray, np.ndarray]:
    key = (fp_name, dim, mem_dtype)
    cached = mem_cache.get(key)
    if cached is not None and not rebuild:
        return cached

    fp_path = fp_cache_path(fp_dir, fp_name, dim)

    if (not rebuild) and fp_path.exists():
        X_u8, vmask = load_fp_from_disk(fp_path)
        if X_u8.ndim != 2 or X_u8.shape[1] != dim or X_u8.shape[0] != len(smiles_list):
            X_u8, vmask = compute_fingerprint_matrix_uint8(smiles_list, fp_name, dim)
            save_fp_to_disk(fp_path, X_u8, vmask)

        X_mem = X_u8.astype(np.float32) if mem_dtype == "float32" else X_u8
        mem_cache.put(key, (X_mem, vmask))
        return X_mem, vmask

    X_u8, vmask = compute_fingerprint_matrix_uint8(smiles_list, fp_name, dim)
    save_fp_to_disk(fp_path, X_u8, vmask)

    X_mem = X_u8.astype(np.float32) if mem_dtype == "float32" else X_u8
    mem_cache.put(key, (X_mem, vmask))
    return X_mem, vmask


# =========================
# 모델 경로 해석(RESULTS -> BACKUP fallback)
# =========================
def resolve_model_path(
    seed_results_dir: Path,
    seed_backup_dir: Path,
    assay: str,
    model_name: str,
    fp: str,
) -> Tuple[Optional[Path], str, Path, Path]:
    """
    반환:
      (used_path or None, source_label, primary_path, backup_path)
    """
    primary = seed_results_dir / f"{assay}_{model_name}_{fp}" / "model.joblib"
    backup = seed_backup_dir / f"{assay}_{model_name}_{fp}" / "model.joblib"

    if primary.exists():
        return primary, "results", primary, backup
    if backup.exists():
        return backup, "backup", primary, backup
    return None, "missing", primary, backup


# =========================
# 메인
# =========================
def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--seed", type=int, required=True, help="seed 번호")
    p.add_argument("--split", type=str, required=True, choices=["train", "val"], help="train 또는 val")
    p.add_argument("--threshold", type=float, default=0.5, help="임계값 (default=0.5)")
    p.add_argument("--default_nbits", type=int, default=2048, help="n_features_in_ 없을 때 fallback")
    p.add_argument("--fp_mem_cache_max", type=int, default=4, help="메모리 FP 캐시 최대 엔트리 수")
    p.add_argument("--fp_mem_dtype", type=str, default="float32", choices=["float32", "uint8"],
                   help="메모리에서 FP dtype")
    p.add_argument("--rebuild_fps", action="store_true", help="fingerprints 캐시 있어도 재생성")
    p.add_argument("--dry-run", action="store_true", help="경로/모델 존재만 점검")
    return p.parse_args()


def main():
    args = parse_args()
    seed = args.seed
    split = args.split
    thr = float(args.threshold)
    fallback_nbits = int(args.default_nbits)

    pred_dir = PRED_BASE_ROOT / f"DART_MTL_seed{seed}{split}"
    input_xlsx = pred_dir / f"seed{seed}{split}.xlsx"
    results_root = pred_dir / "results"
    score_xlsx = pred_dir / f"score_seed{seed}{split}.xlsx"
    fp_dir = pred_dir / "fingerprints"

    train_seed_dir = TRAIN_RESULTS_ROOT / f"seed_{seed}"
    backup_seed_dir = TRAIN_BACKUP_ROOT / f"seed_{seed}"

    print(f"[INFO] input_xlsx      = {input_xlsx}")
    print(f"[INFO] train_seeddir   = {train_seed_dir}")
    print(f"[INFO] backup_seeddir  = {backup_seed_dir}")
    print(f"[INFO] results_root    = {results_root}")
    print(f"[INFO] fp_dir          = {fp_dir}")
    print(f"[INFO] score_xlsx      = {score_xlsx}")
    print(f"[INFO] threshold       = {thr}")
    print(f"[INFO] rebuild_fps     = {args.rebuild_fps}")
    print(f"[INFO] fp_mem_dtype    = {args.fp_mem_dtype}")

    if not input_xlsx.exists():
        raise FileNotFoundError(f"입력 엑셀 없음: {input_xlsx}")
    if not train_seed_dir.exists():
        raise FileNotFoundError(f"학습 seed 디렉토리 없음: {train_seed_dir}")
    # backup_seed_dir는 없어도 되지만, 있으면 fallback 사용

    # 원본 data 로드
    data_df = pd.read_excel(input_xlsx, sheet_name="data", engine=EXCEL_ENGINE)
    smiles_col = find_smiles_col(list(data_df.columns))
    assay_cols = [c for c in data_df.columns if c != smiles_col]
    if not assay_cols:
        raise ValueError("data 시트에 assay 열이 없습니다 (SMILES 외 열이 없음).")

    smiles_list = data_df[smiles_col].tolist()

    if args.dry_run:
        missing = 0
        total = 0
        used_results = 0
        used_backup = 0

        for assay in assay_cols:
            assay_str = str(assay)
            for model_name in MODELS:
                for fp in FINGERPRINTS:
                    total += 1
                    used_path, src, _, _ = resolve_model_path(
                        train_seed_dir, backup_seed_dir, assay_str, model_name, fp
                    )
                    if used_path is None:
                        missing += 1
                    elif src == "results":
                        used_results += 1
                    elif src == "backup":
                        used_backup += 1

        print(f"[DRY-RUN] total={total}, missing_models={missing}, used_results={used_results}, used_backup={used_backup}")
        return

    results_root.mkdir(parents=True, exist_ok=True)
    fp_dir.mkdir(parents=True, exist_ok=True)

    mem_cache = LRUCache(max_items=args.fp_mem_cache_max)
    score_rows = []

    # 조합별(model, fp) → 엑셀 1개
    for model_name in MODELS:
        for fp in FINGERPRINTS:
            combo_dir = results_root / f"{model_name}_{fp}"
            combo_dir.mkdir(parents=True, exist_ok=True)
            out_xlsx = combo_dir / f"seed{seed}{split}_{model_name}_{fp}.xlsx"

            # 워크북 새로 작성(원본 템플릿 복사한 다음 시트 교체)
            shutil.copy2(input_xlsx, out_xlsx)
            wb = openpyxl.load_workbook(out_xlsx)

            # 원본 정답 보존
            write_df_sheet(wb, data_df, "truth_data")

            # 결과용 DataFrame 틀 준비 (SMILES + assay 전체)
            prob_df = data_df.copy()
            pred_df = data_df.copy()

            # 이 조합의 pred_info
            info_rows = []
            missing_models = 0

            for assay in assay_cols:
                assay_str = str(assay)

                used_path, src_label, primary_path, backup_path = resolve_model_path(
                    train_seed_dir, backup_seed_dir, assay_str, model_name, fp
                )

                if used_path is None:
                    missing_models += 1
                    prob_df[assay_str] = np.nan
                    pred_df[assay_str] = np.nan
                    info_rows.append({
                        "assay": assay_str,
                        "model": model_name,
                        "fp": fp,
                        "status": "missing_model",
                        "model_source": "missing",
                        "model_path_used": None,
                        "primary_path": str(primary_path),
                        "backup_path": str(backup_path),
                        "dim_used": None,
                        "note": "not found in results or backup",
                    })
                    score_rows.append({
                        "Assay Name": assay_str,
                        "Model": model_name,
                        "MF Metric": fp,
                        "AUC": np.nan,
                        "accuracy": np.nan,
                        "precision": np.nan,
                        "recall": np.nan,
                        "F1": np.nan,
                        "n_eval": 0,
                    })
                    continue

                # 모델 로드
                try:
                    model = joblib_load(used_path)
                except Exception as e:
                    prob_df[assay_str] = np.nan
                    pred_df[assay_str] = np.nan
                    info_rows.append({
                        "assay": assay_str,
                        "model": model_name,
                        "fp": fp,
                        "status": "load_failed",
                        "model_source": src_label,
                        "model_path_used": str(used_path),
                        "primary_path": str(primary_path),
                        "backup_path": str(backup_path),
                        "dim_used": None,
                        "note": str(e),
                    })
                    score_rows.append({
                        "Assay Name": assay_str,
                        "Model": model_name,
                        "MF Metric": fp,
                        "AUC": np.nan,
                        "accuracy": np.nan,
                        "precision": np.nan,
                        "recall": np.nan,
                        "F1": np.nan,
                        "n_eval": 0,
                    })
                    continue

                # dim 결정 + FP 로드/생성
                if fp == "MACCS":
                    dim = MACCS_BITS
                    nfeat = get_n_features_in(model)
                    if nfeat is not None and nfeat != MACCS_BITS:
                        prob_df[assay_str] = np.nan
                        pred_df[assay_str] = np.nan
                        info_rows.append({
                            "assay": assay_str,
                            "model": model_name,
                            "fp": fp,
                            "status": "dim_mismatch_maccs",
                            "model_source": src_label,
                            "model_path_used": str(used_path),
                            "primary_path": str(primary_path),
                            "backup_path": str(backup_path),
                            "dim_used": dim,
                            "note": f"model expects {nfeat}, MACCS fixed {MACCS_BITS}",
                        })
                        score_rows.append({
                            "Assay Name": assay_str,
                            "Model": model_name,
                            "MF Metric": fp,
                            "AUC": np.nan,
                            "accuracy": np.nan,
                            "precision": np.nan,
                            "recall": np.nan,
                            "F1": np.nan,
                            "n_eval": 0,
                        })
                        continue
                else:
                    dim = get_n_features_in(model) or fallback_nbits

                X_mem, vmask = load_or_build_fp(
                    smiles_list=smiles_list,
                    fp_name=fp,
                    dim=dim,
                    fp_dir=fp_dir,
                    mem_cache=mem_cache,
                    mem_dtype=args.fp_mem_dtype,
                    rebuild=args.rebuild_fps,
                )

                # 예측
                try:
                    X_in = ensure_feature_named_input(model, X_mem)
                    y_proba_all = safe_predict_proba(model, X_in).astype(np.float64)
                except Exception as e:
                    prob_df[assay_str] = np.nan
                    pred_df[assay_str] = np.nan
                    info_rows.append({
                        "assay": assay_str,
                        "model": model_name,
                        "fp": fp,
                        "status": "predict_failed",
                        "model_source": src_label,
                        "model_path_used": str(used_path),
                        "primary_path": str(primary_path),
                        "backup_path": str(backup_path),
                        "dim_used": dim,
                        "note": str(e),
                    })
                    score_rows.append({
                        "Assay Name": assay_str,
                        "Model": model_name,
                        "MF Metric": fp,
                        "AUC": np.nan,
                        "accuracy": np.nan,
                        "precision": np.nan,
                        "recall": np.nan,
                        "F1": np.nan,
                        "n_eval": 0,
                    })
                    continue

                if y_proba_all.shape[0] != len(data_df):
                    prob_df[assay_str] = np.nan
                    pred_df[assay_str] = np.nan
                    info_rows.append({
                        "assay": assay_str,
                        "model": model_name,
                        "fp": fp,
                        "status": "length_mismatch",
                        "model_source": src_label,
                        "model_path_used": str(used_path),
                        "primary_path": str(primary_path),
                        "backup_path": str(backup_path),
                        "dim_used": dim,
                        "note": f"pred_len={y_proba_all.shape[0]} vs N={len(data_df)}",
                    })
                    score_rows.append({
                        "Assay Name": assay_str,
                        "Model": model_name,
                        "MF Metric": fp,
                        "AUC": np.nan,
                        "accuracy": np.nan,
                        "precision": np.nan,
                        "recall": np.nan,
                        "F1": np.nan,
                        "n_eval": 0,
                    })
                    continue

                # invalid SMILES는 NaN 처리
                y_proba_all[~vmask] = np.nan
                prob_df[assay_str] = y_proba_all

                # pred(0/1) 만들기: NaN은 그대로 NaN
                y_pred_all = np.where(np.isnan(y_proba_all), np.nan, (y_proba_all >= thr).astype(int))
                pred_df[assay_str] = y_pred_all

                info_rows.append({
                    "assay": assay_str,
                    "model": model_name,
                    "fp": fp,
                    "status": "ok",
                    "model_source": src_label,
                    "model_path_used": str(used_path),
                    "primary_path": str(primary_path),
                    "backup_path": str(backup_path),
                    "dim_used": dim,
                    "note": "used results model" if src_label == "results" else "used backup model",
                })

                # 점수 계산: 정답 0/1만 + 예측 NaN 제외
                y_raw = data_df[assay_str]
                y_num = pd.to_numeric(y_raw, errors="coerce")
                labeled_mask = y_num.isin([0, 1]).to_numpy()
                usable_mask = labeled_mask & (~np.isnan(y_proba_all))
                n_eval = int(usable_mask.sum())

                if n_eval > 0:
                    y_true = y_num.to_numpy()[usable_mask].astype(int)
                    y_proba = y_proba_all[usable_mask]
                    metrics = compute_metrics(y_true, y_proba, thr=thr)
                else:
                    metrics = {"AUC": np.nan, "accuracy": np.nan, "precision": np.nan, "recall": np.nan, "F1": np.nan}

                score_rows.append({
                    "Assay Name": assay_str,
                    "Model": model_name,
                    "MF Metric": fp,
                    "AUC": metrics["AUC"],
                    "accuracy": metrics["accuracy"],
                    "precision": metrics["precision"],
                    "recall": metrics["recall"],
                    "F1": metrics["F1"],
                    "n_eval": n_eval,
                })

            # data_prob / data_pred 시트 작성
            write_df_sheet(wb, prob_df, "data_prob")
            write_df_sheet(wb, pred_df, "data_pred")
            write_pred_info_sheet(wb, info_rows, "pred_info")

            # 기존 원본 data 시트는 혼동 방지로 제거(원하면 유지해도 됨)
            if "data" in wb.sheetnames:
                wb.remove(wb["data"])

            wb.save(out_xlsx)
            wb.close()

            print(f"[OK] combo={model_name}_{fp} | missing_models={missing_models} -> {out_xlsx}")

    # score 저장
    score_df = pd.DataFrame(score_rows, columns=[
        "Assay Name", "Model", "MF Metric", "AUC", "accuracy", "precision", "recall", "F1", "n_eval"
    ])
    score_df.to_excel(score_xlsx, index=False, engine=EXCEL_ENGINE)
    print(f"\n[SAVED] score file: {score_xlsx}")


if __name__ == "__main__":
    main()
