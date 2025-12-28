#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
merge_dart_mtl_scores.py

요구사항
1) seed, train/val로 분리된 score 엑셀들을 하나의 엑셀로 합치기
   예) .../DART_MTL_seed0train/score_seed0train.xlsx
       .../DART_MTL_seed0val/score_seed0val.xlsx
   -> 출력 엑셀에서 시트명: seed0train, seed0val, seed1train, ...

2) 추가 시트: total_train, total_val
   - 컬럼: assay_name, model, mf, AUC, accuracy, precision, recall, F1 (+ n_seed)
   - 각 (assay_name, model, mf) 조합에 대해 seed 0~2의 성능을 모아서
     평균(±표준편차)을 문자열로 기록 (표준편차는 ddof=0 기준)
   - 같은 assay_name끼리 연속되도록 정렬하고,
     total_* 시트에서 assay_name 열은 같은 값 구간을 "셀 병합"하여 표시

3) 추가 시트: best_train, best_val
   - total_*과 동일한 양식 + assay_name 병합 유지
   - 단, "model은 전부 유지" (assay_name, model 조합마다 1행)
   - (assay_name, model)별로 F1 평균이 가장 높은 fp(mf) 1개만 선택해서 기입
     => 같은 assay_name에서 모델별로 best fp가 달라도 OK

추가 요구:
- total_*/best_*에서 model 나열 순서: RF, XGB, LR, GBT, DT

참고:
- openpyxl merge_cells: https://openpyxl.readthedocs.io/en/stable/usage.html#merging-unmerging-cells
- openpyxl dataframe_to_rows: https://openpyxl.readthedocs.io/en/stable/pandas.html
"""

import argparse
import re
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font


# -------------------------
# 기존 score 엑셀 기준 컬럼
# -------------------------
COL_ASSAY = "Assay Name"
COL_MODEL = "Model"
COL_MF = "MF Metric"
METRICS = ["AUC", "accuracy", "precision", "recall", "F1"]

# -------------------------
# 모델 출력 순서 지정 (요구)
# score 파일의 Model 값이 보통: dt, rf, xgb, gbt, logistic
# 이를 RF, XGB, LR, GBT, DT 순으로 정렬
# -------------------------
MODEL_ORDER = ["rf", "xgb", "logistic", "gbt", "dt"]
MODEL_RANK = {m: i for i, m in enumerate(MODEL_ORDER)}


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument(
        "--base_dir",
        type=str,
        required=True,
        help="DART_MTL_251228 디렉토리 (예: .../experiments/prediction/DART_MTL_251228)",
    )
    p.add_argument(
        "--out_xlsx",
        type=str,
        required=True,
        help="출력 엑셀 경로 (예: .../merged_scores.xlsx)",
    )
    p.add_argument(
        "--seeds",
        type=str,
        default="0,1,2",
        help="사용할 seed 목록 (기본: 0,1,2)",
    )
    p.add_argument(
        "--splits",
        type=str,
        default="train,val",
        help="사용할 split 목록 (기본: train,val)",
    )
    return p.parse_args()


def normalize_seed_split_lists(seeds_str: str, splits_str: str) -> Tuple[List[int], List[str]]:
    seeds = []
    for s in seeds_str.split(","):
        s = s.strip()
        if s:
            seeds.append(int(s))

    splits = []
    for sp in splits_str.split(","):
        sp = sp.strip()
        if sp:
            splits.append(sp)

    return seeds, splits


def find_score_files(base_dir: Path, seeds: List[int], splits: List[str]) -> Dict[Tuple[int, str], Path]:
    """
    기대 구조:
      DART_MTL_seed{seed}{split}/score_seed{seed}{split}.xlsx
    """
    found: Dict[Tuple[int, str], Path] = {}
    for seed in seeds:
        for split in splits:
            d = base_dir / f"DART_MTL_seed{seed}{split}"
            f = d / f"score_seed{seed}{split}.xlsx"
            if f.exists():
                found[(seed, split)] = f

    # 보조 탐색(약간의 변형 대비)
    if len(found) == 0:
        pat_dir = re.compile(r"^DART_MTL_seed(\d+)(train|val)$", re.IGNORECASE)
        pat_file = re.compile(r"^score_seed(\d+)(train|val)\.xlsx$", re.IGNORECASE)
        for d in base_dir.iterdir():
            if not d.is_dir():
                continue
            m = pat_dir.match(d.name)
            if not m:
                continue
            seed = int(m.group(1))
            split = m.group(2).lower()
            if seed not in seeds or split not in splits:
                continue
            for f in d.iterdir():
                if f.is_file() and pat_file.match(f.name):
                    found[(seed, split)] = f
                    break

    return found


def read_score_xlsx(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, engine="openpyxl")
    needed = [COL_ASSAY, COL_MODEL, COL_MF] + METRICS
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise ValueError(
            f"score 파일 컬럼이 예상과 다름: {path}\nmissing={missing}\ncols={list(df.columns)}"
        )
    for m in METRICS:
        df[m] = pd.to_numeric(df[m], errors="coerce")
    return df


def safe_sheet_name(name: str) -> str:
    name = str(name)
    return name if len(name) <= 31 else name[:31]


def write_df_to_sheet(wb: openpyxl.Workbook, sheet_name: str, df: pd.DataFrame):
    sheet_name = safe_sheet_name(sheet_name)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    header_font = Font(bold=True)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c_idx in range(1, len(row) + 1):
                ws.cell(row=1, column=c_idx).font = header_font

    ws.freeze_panes = "A2"
    return ws


def autosize_columns(ws: openpyxl.worksheet.worksheet.Worksheet, max_width: int = 55):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _fmt_mean_std(vals: np.ndarray) -> str:
    vals = vals[~np.isnan(vals)]
    if vals.size == 0:
        return ""
    mean = float(np.mean(vals))
    std = float(np.std(vals, ddof=0))  # 요구: ddof=0
    return f"{mean:.6f}(±{std:.6f})"


def _model_rank_series(model_series: pd.Series) -> pd.Series:
    """
    정렬용 랭크: rf < xgb < logistic < gbt < dt
    모르는 모델은 맨 뒤로.
    """
    s = model_series.astype(str).str.strip().str.lower()
    return s.map(MODEL_RANK).fillna(len(MODEL_ORDER)).astype(int)


def summarize_split_mean_std(dfs_by_seed: Dict[int, pd.DataFrame]) -> pd.DataFrame:
    """
    seed별 score df를 합쳐서 (assay_name, model, mf)별 mean±std 문자열 생성
    + best 선택을 위해 _F1_mean_num(숫자) 포함
    """
    rows = []
    for seed, df in dfs_by_seed.items():
        tmp = df.copy()
        tmp["seed"] = seed
        rows.append(tmp)

    if not rows:
        return pd.DataFrame(columns=["assay_name", "model", "mf"] + METRICS + ["n_seed", "_F1_mean_num", "_model_rank"])

    all_df = pd.concat(rows, ignore_index=True).rename(
        columns={COL_ASSAY: "assay_name", COL_MODEL: "model", COL_MF: "mf"}
    )

    out_rows = []
    keys = ["assay_name", "model", "mf"]

    for (assay, model, mf), g in all_df.groupby(keys, dropna=False):
        r = {"assay_name": assay, "model": model, "mf": mf}
        r["n_seed"] = int(g["seed"].nunique())

        for metric in METRICS:
            vals = pd.to_numeric(g[metric], errors="coerce").to_numpy()
            r[metric] = _fmt_mean_std(vals)

        f1_vals = pd.to_numeric(g["F1"], errors="coerce").to_numpy()
        f1_vals = f1_vals[~np.isnan(f1_vals)]
        r["_F1_mean_num"] = float(np.mean(f1_vals)) if f1_vals.size > 0 else np.nan

        # 모델 정렬 랭크
        r["_model_rank"] = int(MODEL_RANK.get(str(model).strip().lower(), len(MODEL_ORDER)))

        out_rows.append(r)

    out_df = pd.DataFrame(out_rows)

    # 보기 좋게 정렬: assay_name -> model_order -> mf
    out_df = out_df.sort_values(
        ["assay_name", "_model_rank", "model", "mf"],
        kind="mergesort",
    ).reset_index(drop=True)
    return out_df


def best_fp_per_assay_model(total_raw: pd.DataFrame) -> pd.DataFrame:
    """
    (assay_name, model)별로 best fp(mf)를 1개 선택:
      - 기준: _F1_mean_num 최대
      - 동점 처리: mf 사전순(오름차순)으로 가장 먼저 나오는 것을 선택(안정정렬)
    반환: (assay_name, model)당 1행 (즉 model은 전부 유지)
    """
    if total_raw.empty:
        return total_raw.copy()

    tmp = total_raw.copy()
    tmp["_F1_mean_num"] = pd.to_numeric(tmp["_F1_mean_num"], errors="coerce")

    # 정렬: assay -> model_order -> F1mean desc -> mf asc
    tmp = tmp.sort_values(
        ["assay_name", "_model_rank", "model", "_F1_mean_num", "mf"],
        ascending=[True, True, True, False, True],
        kind="mergesort",
    )

    best = tmp.groupby(["assay_name", "model"], as_index=False).head(1).reset_index(drop=True)

    # 최종 정렬: assay -> model_order
    best = best.sort_values(
        ["assay_name", "_model_rank", "model"],
        kind="mergesort",
    ).reset_index(drop=True)
    return best


def merge_assay_name_cells(ws: openpyxl.worksheet.worksheet.Worksheet, assay_col: int = 1):
    """
    같은 assay_name이 연속된 구간을 병합
    """
    max_row = ws.max_row
    if max_row <= 2:
        return

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    start = 2
    cur = ws.cell(row=2, column=assay_col).value

    for r in range(3, max_row + 2):  # sentinel
        v = ws.cell(row=r, column=assay_col).value if r <= max_row else None
        if v != cur:
            end = r - 1
            if end > start:
                ws.merge_cells(
                    start_row=start, start_column=assay_col,
                    end_row=end, end_column=assay_col
                )
                ws.cell(row=start, column=assay_col).alignment = center
            else:
                ws.cell(row=start, column=assay_col).alignment = center
            start = r
            cur = v

    # 전체 데이터 영역 세로 중앙 정렬
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(
                horizontal=(cell.alignment.horizontal if cell.alignment else None),
                vertical="center",
                wrap_text=True,
            )


def main():
    args = parse_args()
    base_dir = Path(args.base_dir)
    out_xlsx = Path(args.out_xlsx)
    seeds, splits = normalize_seed_split_lists(args.seeds, args.splits)

    if not base_dir.exists():
        raise FileNotFoundError(f"base_dir 없음: {base_dir}")

    found = find_score_files(base_dir, seeds=seeds, splits=splits)
    if not found:
        raise FileNotFoundError(
            "score 파일을 하나도 찾지 못했습니다.\n"
            f"base_dir={base_dir}\n"
            f"expected like: DART_MTL_seed0train/score_seed0train.xlsx"
        )

    print(f"[INFO] base_dir = {base_dir}")
    print(f"[INFO] out_xlsx = {out_xlsx}")
    print(f"[INFO] seeds    = {seeds}")
    print(f"[INFO] splits   = {splits}")
    print(f"[INFO] found score files = {len(found)}")

    # split별로 seed->df 저장
    split_seed_dfs: Dict[str, Dict[int, pd.DataFrame]] = {sp: {} for sp in splits}

    # 워크북 생성
    wb = openpyxl.Workbook()
    if wb.sheetnames:
        wb.remove(wb[wb.sheetnames[0]])

    # 1) seed/split 개별 시트 (원본 순서는 유지)
    for seed in seeds:
        for split in splits:
            p = found.get((seed, split))
            if p is None:
                print(f"[WARN] missing: seed{seed}{split} score file")
                continue

            df = read_score_xlsx(p)
            split_seed_dfs[split][seed] = df

            sheet_name = f"seed{seed}{split}"
            ws = write_df_to_sheet(wb, sheet_name, df)
            autosize_columns(ws)
            print(f"[OK] wrote sheet: {sheet_name} <- {p}")

    # 공통 컬럼 순서(출력용)
    col_order = ["assay_name", "model", "mf"] + METRICS + ["n_seed"]

    # 2) total_* + best_* 시트
    for split in splits:
        total_raw = summarize_split_mean_std(split_seed_dfs[split])

        # total_* 출력용 df
        total_df = total_raw.copy()
        for c in col_order:
            if c not in total_df.columns:
                total_df[c] = ""
        total_df = total_df[col_order].reset_index(drop=True)

        ws_total = write_df_to_sheet(wb, f"total_{split}", total_df)
        merge_assay_name_cells(ws_total, assay_col=1)
        autosize_columns(ws_total)
        print(f"[OK] wrote sheet: total_{split} (rows={len(total_df)})")

        # best_*: (assay, model)별 best fp 1개 선택 (model 전체 유지)
        best_raw = best_fp_per_assay_model(total_raw)
        best_df = best_raw.copy()
        for c in col_order:
            if c not in best_df.columns:
                best_df[c] = ""
        best_df = best_df[col_order].reset_index(drop=True)

        ws_best = write_df_to_sheet(wb, f"best_{split}", best_df)
        merge_assay_name_cells(ws_best, assay_col=1)
        autosize_columns(ws_best)
        print(f"[OK] wrote sheet: best_{split} (rows={len(best_df)})")

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    wb.close()
    print(f"\n[SAVED] {out_xlsx}")


if __name__ == "__main__":
    main()
